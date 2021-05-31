import openpyxl
import matplotlib.pyplot as plt
import numpy as np

bk = openpyxl.load_workbook('机器人引论.xlsx')
sheet = bk.active

minrow = sheet.min_row  # 最小行
maxrow = sheet.max_row  # 最大行
mincol = sheet.min_column  # 最小列
maxcol = sheet.max_column  # 最大列

"""
#按行读取
for i in range(minrow, maxrow + 1):
	for j in range(mincol, maxcol + 1):
		cell = sheet.cell(i, j).value
		print(cell, end=" ")
	print()
"""

"""
#按列读取
for m in range(mincol, maxcol + 1):
	for n in range(minrow, maxrow + 1):
		cell = sheet.cell(n, m).value
		print(cell, end=" ")
	print()
"""

"""
# 访问A1至C3范围单元格
cell_range = ws2['A1':'C3']
# 访问A列所有存在数据的单元格
colA = ws2['A']
# 访问A列到C列所有存在数据的单元格
col_range = ws2['A:C']
# 访问第1行所有存在数据的单元格
row1 = ws2[1]
# 访问第1行至第5行所有存在数据的单元格
row_range = ws2[1:5]
"""

# from 1 to 64
id = [sheet['A'][i].value for i in range(len(sheet['A']))]
gender = [sheet['G'][i].value for i in range(len(sheet['G']))]
age = [sheet['H'][i].value for i in range(len(sheet['H']))]
degree = [sheet['Q'][i].value for i in range(len(sheet['Q']))]
print(degree)

# question 23
ques_23_job = [sheet['AU'][i].value for i in range(len(sheet['AU']))]
ques_23_ethic = [sheet['AV'][i].value for i in range(len(sheet['AV']))]
ques_23_industry = [sheet['AW'][i].value for i in range(len(sheet['AW']))]
ques_23_self = [sheet['AX'][i].value for i in range(len(sheet['AX']))]
ques_23_relation = [sheet['AY'][i].value for i in range(len(sheet['AY']))]

# question 22
ques_22 = [sheet['AT'][i].value for i in range(len(sheet['AT']))]

# question 21
ques_21 = [sheet['AS'][i].value for i in range(len(sheet['AS']))]

# question 20
ques_20 = [sheet['AR'][i].value for i in range(len(sheet['AR']))]

# question 19
ques_19 = [sheet['AQ'][i].value for i in range(len(sheet['AQ']))]

# question 18
ques_18 = [sheet['AP'][i].value for i in range(len(sheet['AP']))]

# question 17
ques_17 = [sheet['AO'][i].value for i in range(len(sheet['AO']))]

# question 16
ques_16 = [sheet['AN'][i].value for i in range(len(sheet['AN']))]

# question 15
ques_15 = [sheet['AM'][i].value for i in range(len(sheet['AM']))]

data_men = [0, 0]
data_wom = [0, 0]

data_age1 = [0, 0]
data_age2 = [0, 0]
data_age3 = [0, 0]
data_age4 = [0, 0]

data_degree1 = [0, 0]
data_degree2 = [0, 0]
data_degree3 = [0, 0]
data_degree4 = [0, 0]


def addDegree(id, judge):
    if degree[id] == 1:
        data_degree1[judge] += 1
    elif degree[id] == 2:
        data_degree2[judge] += 1
    elif degree[id] == 3:
        data_degree3[judge] += 1
    else:
        data_degree4[judge] += 1


def addGender(id, judge):
    if gender[id] == 1:
        data_men[judge] += 1
    else:
        data_wom[judge] += 1


def addAge(id, judge):
    if age[id] == 1:
        data_age1[judge] += 1
    elif age[id] == 2:
        data_age2[judge] += 1
    elif age[id] == 3:
        data_age3[judge] += 1
    else:
        data_age4[judge] += 1


description = 'Question %d' % 17

for i in range(len(id)):
    this_value = ques_17[i]
    if i == 0:
        print(this_value)
        continue
    if this_value == -3:
        continue
    elif this_value == 1:
        addGender(i, 1)
        addAge(i, 1)
        addDegree(i, 1)
    else:
        addGender(i, 0)
        addAge(i, 0)
        addDegree(i, 0)

labels = ['men', 'women']
no_s = [data_men[0], data_wom[0]]
yes_s = [data_men[1], data_wom[1]]

x = np.arange(len(labels))  # the label locations
width = 0.35  # the width of the bars

fig, ax = plt.subplots()
rects1 = ax.bar(x - width / 2, no_s, width, label='No')
rects2 = ax.bar(x + width / 2, yes_s, width, label='Yes')

# Add some text for labels, title and custom x-axis tick labels, etc.
ax.set_ylabel('person cnt')
ax.set_title(description)
ax.set_xticks(x)
ax.set_xticklabels(labels)
ax.legend()

ax.bar_label(rects1, padding=3)
ax.bar_label(rects2, padding=3)

fig.tight_layout()

plt.show()

labels = ['5-16', '17-28', '29-50', '50-']
no_s = [data_age1[0], data_age2[0], data_age3[0], data_age4[0]]
yes_s = [data_age1[1], data_age2[1], data_age3[1], data_age4[1]]

x = np.arange(len(labels))  # the label locations
width = 0.35  # the width of the bars

fig, ax = plt.subplots()
rects1 = ax.bar(x - width / 2, no_s, width, label='No')
rects2 = ax.bar(x + width / 2, yes_s, width, label='Yes')

# Add some text for labels, title and custom x-axis tick labels, etc.
ax.set_ylabel('person cnt')
ax.set_title(description)
ax.set_xticks(x)
ax.set_xticklabels(labels)
ax.legend()

ax.bar_label(rects1, padding=3)
ax.bar_label(rects2, padding=3)

fig.tight_layout()

plt.show()


labels = ['Work in', 'Major in', 'Interested', 'Boring']
no_s = [data_degree1[0], data_degree2[0], data_degree3[0], data_degree4[0]]
yes_s = [data_degree1[1], data_degree2[1], data_degree3[1], data_degree4[1]]

x = np.arange(len(labels))  # the label locations
width = 0.35  # the width of the bars

fig, ax = plt.subplots()
rects1 = ax.bar(x - width / 2, no_s, width, label='No')
rects2 = ax.bar(x + width / 2, yes_s, width, label='Yes')

# Add some text for labels, title and custom x-axis tick labels, etc.
ax.set_ylabel('person cnt')
ax.set_title(description)
ax.set_xticks(x)
ax.set_xticklabels(labels)
ax.legend()

ax.bar_label(rects1, padding=3)
ax.bar_label(rects2, padding=3)

fig.tight_layout()

plt.show()